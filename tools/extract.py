# -*- coding: utf-8 -*-
"""朝代跨度表.xlsx -> docs/data.json 解析脚本

用法:  uv run --with openpyxl tools/extract.py

原表结构:
  - Sheet4, 行 2..410, 每行 10 年, 行 2 = 公元前 2030 年
  - 列 B..AQ 共 42 列, 分属 7 大区
  - 政权 = 带名字的合并单元格「核心区」+ 周边同色散落格 (不规则区域)

产物:
  - docs/data.json          页面数据
  - tools/report.md        解析/修正审核报告
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "朝代跨度表.xlsx"
SHEET = "Sheet4"

MIN_ROW, MAX_ROW = 2, 410
MIN_COL, MAX_COL = 2, 43  # B..AQ
YEAR_START, YEAR_STEP = -2030, 10

REGIONS = [
    ("东南亚", 2, 4),
    ("青藏高原", 5, 8),
    ("西域", 9, 12),
    ("中原", 13, 31),
    ("东北", 32, 35),
    ("外蒙古", 36, 39),
    ("中亚", 40, 43),
]

# Office 默认主题色 (theme index 0..9)
THEME_RGB = [
    "FFFFFF", "000000", "E7E6E6", "44546A", "5B9BD5",
    "ED7D31", "A5A5A5", "FFC000", "4472C4", "70AD47",
    "0563C1", "954F72",
]

# ---------------------------------------------------------------------------
# 名称修正 (raw -> 修正后)。改动会记录进 report.md 供审核。
RENAMES = {
    # 错别字 / 规范写法
    "钱缪": "钱镠",
    "鈡傅": "钟传",
    "宝融": "窦融",
    "朱桀": "朱粲",
    "准格尔": "准噶尔蒙古",
    "准格尔蒙古": "准噶尔蒙古",
    "黠嘎斯": "黠戛斯",
    "桂溪": "桂系",
    # 口语化 -> 中性表述
    "小日本入侵": "日占东北",
    "豫系、直系、湘系、桂溪等": "直系、皖系、湘系、桂系等",
    "李宗仁、西北军晋绥军、川军等等": "李宗仁、西北军、晋绥军、川军等",
    "女真族归附": "女真诸部（附明）",
    # 截断名补全
    "王审": "王审知",
    "窦建": "窦建德",
    "刘黑": "刘黑闼",
    "梁师": "梁师都",
    "公孙": "公孙瓒",
    "韩马": "韩遂、马腾",
    # 跨格拆分的名字拼回 (与同色相邻格归组)
    "李茂": "李茂贞",
    "贞": "李茂贞",
    "刘仁": "刘仁恭",
    "恭": "刘仁恭",
    # 政权归并: 括号补注/开国者/阶段标签归入本体
    "（巴氐）": "成汉",
    "半殖民地化": "清",
    "李渊": "唐",
    "刘秀": "东汉",
    "朱元璋": "明",
    "仇池（氐）": "仇池",
    "喀尔喀": "喀尔喀蒙古",
    "建肇": "武泰军",
    "武泰": "武泰军",
    # 占领事件标签, 补全语义
    "齐占": "齐占燕",
    "五国": "五国伐齐",
    # 顿号分组 (原表竖排换行分组, 拉平后补顿号)
    "义渠西戎等": "义渠、西戎等",
    "巴蜀等": "巴、蜀等",
    "赤狄白狄": "赤狄、白狄",
    "郑卫等": "郑、卫等",
    "贾杨等": "贾、杨等",
    "唐随等": "唐、随等",
    "南越夜郎等": "南越、夜郎等",
    "唐旄发羌象雄宝髻": "唐旄、发羌、象雄、宝髻",
    "党项女国等部": "党项、女国等部",
    # 消歧
    "天德": "天德军",
}

# 按坐标定点改名 (同名不同政权无法用 RENAMES 区分)
RENAME_AT = {
    "AA340": "宋（红巾）",
    "Y341": "周（张士诚）",
    "D342": "梁王（元）",
    "AC296": "燕（刘守光）",
    "M239": "代",    # （鲜卑）补注: 拓跋代国
    "T241": "前燕",  # （鲜卑）补注: 慕容前燕
}

# 人工补充的标签区: (r0, c0, r1, c1, name, virtual)
# virtual=True 表示只作切分锚点 (著名分界年份/无标签色块归名), 不渲染标签、
# 若与任何色块无交集则静默丢弃
EXTRA_SEEDS = [
    (397, 5, 399, 8, "西藏地方（噶厦）", False),  # 民国时期青藏, 原表留白
    (266, 8, 266, 8, "隋", True),        # 隋灭吐谷浑设四郡 (609)
    (267, 8, 271, 8, "吐谷浑", True),    # 隋末吐谷浑复国
    (267, 11, 270, 11, "吐谷浑", True),
    (300, 30, 304, 31, "辽", True),      # 契丹 947 改国号辽
    (318, 17, 321, 21, "南宋", True),    # 靖康之变 1127
    (332, 20, 334, 24, "元", True),      # 忽必烈 1271 建元
]

# 近似重色归一 (原表笔误用了两个几乎相同的颜色)
COLOR_ALIAS = {
    "#9DC3E6": "#9BC2E6",  # 曹魏东半部 / 西域长史府
    "#DEEBF7": "#DDEBF7",  # 前燕 / 后燕 / 北燕
}

# ---------------------------------------------------------------------------


def apply_tint(hex6: str, tint: float) -> str:
    out = []
    for i in (0, 2, 4):
        v = int(hex6[i:i + 2], 16)
        if tint < 0:
            v = v * (1 + tint)
        else:
            v = v * (1 - tint) + 255 * tint
        out.append(max(0, min(255, round(v))))
    return "".join(f"{v:02X}" for v in out)


def resolve_fill(fill):
    """PatternFill -> '#RRGGBB' 或 None"""
    if fill is None or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb":
        rgb = fg.rgb
        if not rgb or rgb == "00000000":
            return None
        return "#" + rgb[-6:]
    if fg.type == "theme":
        try:
            base = THEME_RGB[fg.theme]
        except (IndexError, TypeError):
            return None
        return "#" + apply_tint(base, fg.tint or 0.0)
    return None


def clean_name(v) -> str:
    s = re.sub(r"[\s　]+", "", str(v))
    return s.replace("(", "（").replace(")", "）")


def year_of(row: int) -> int:
    return YEAR_START + (row - MIN_ROW) * YEAR_STEP


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # --- 合并单元格索引 ---
    anchor_of = {}  # (r,c) -> (r0,c0)
    merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row < MIN_ROW or mr.min_col < MIN_COL or mr.min_col > MAX_COL:
            continue
        merges.append(mr)
        for r in range(mr.min_row, min(mr.max_row, MAX_ROW) + 1):
            for c in range(mr.min_col, min(mr.max_col, MAX_COL) + 1):
                anchor_of[(r, c)] = (mr.min_row, mr.min_col)

    # --- 颜色网格 (合并区内空样式的格子继承锚点颜色) ---
    color = {}
    for r in range(MIN_ROW, MAX_ROW + 1):
        for c in range(MIN_COL, MAX_COL + 1):
            col = resolve_fill(ws.cell(r, c).fill)
            if col is None and (r, c) in anchor_of:
                col = resolve_fill(ws.cell(*anchor_of[(r, c)]).fill)
            if col and col.upper() not in ("#FFFFFF",):
                col = col.upper()
                color[(r, c)] = COLOR_ALIAS.get(col, col)

    # --- 标签种子: 合并区 + 散落带字单元格 + 人工补充 ---
    renames_used = []  # (raw, new, where)
    seeds = []  # dict(rect, name, raw, anchor)

    def add_seed(r0, c0, r1, c1, rawval, coord):
        raw = clean_name(rawval)
        if not raw:
            return
        name = RENAMES.get(raw, raw)
        if coord in RENAME_AT:
            name = RENAME_AT[coord]
        if name != raw:
            renames_used.append((raw, name, coord))
        seeds.append(dict(rect=(r0, c0, r1, c1), name=name, raw=raw,
                          coord=coord, virtual=False))

    for mr in merges:
        v = ws.cell(mr.min_row, mr.min_col).value
        if v is not None:
            add_seed(mr.min_row, mr.min_col, min(mr.max_row, MAX_ROW),
                     min(mr.max_col, MAX_COL), v,
                     f"{get_column_letter(mr.min_col)}{mr.min_row}")
    for r in range(MIN_ROW, MAX_ROW + 1):
        for c in range(MIN_COL, MAX_COL + 1):
            if (r, c) in anchor_of:
                continue
            v = ws.cell(r, c).value
            if v is not None:
                add_seed(r, c, r, c, v, f"{get_column_letter(c)}{r}")
    for (r0, c0, r1, c1, name, virtual) in EXTRA_SEEDS:
        raw = "（切分锚点）" if virtual else "（原表留白）"
        seeds.append(dict(rect=(r0, c0, r1, c1), name=name, raw=raw,
                          coord=f"{get_column_letter(c0)}{r0}+", virtual=virtual))
        renames_used.append((raw, name, f"{get_column_letter(c0)}{r0}"))

    # --- 同色连通块 ---
    comp_of = {}
    comps = []  # list of dict(cells=set, color=str)
    for cell in color:
        if cell in comp_of:
            continue
        cid = len(comps)
        col = color[cell]
        stack = [cell]
        comp_of[cell] = cid
        cells = set()
        while stack:
            cur = stack.pop()
            cells.add(cur)
            r, c = cur
            for nb in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if nb not in comp_of and color.get(nb) == col:
                    comp_of[nb] = cid
                    stack.append(nb)
        comps.append(dict(cells=cells, color=col))

    # --- 种子归属连通块 ---
    def rect_cells(rect):
        r0, c0, r1, c1 = rect
        return [(r, c) for r in range(r0, r1 + 1) for c in range(c0, c1 + 1)]

    comp_seeds = defaultdict(list)  # cid -> seeds
    nofill_seeds = []
    for s in seeds:
        cids = {comp_of[cell] for cell in rect_cells(s["rect"]) if cell in comp_of}
        if cids:
            # 一般只落在一个连通块; 若跨多块取覆盖格数最多的
            best = max(cids, key=lambda cid: sum(
                1 for cell in rect_cells(s["rect"]) if comp_of.get(cell) == cid))
            comp_seeds[best].append(s)
        elif not s["virtual"]:
            nofill_seeds.append(s)  # 无填充色政权 (如 夏 / 商)

    # --- 连通块 -> 政权 (多组标签时按最近标签切分) ---
    def rect_dist(cell, rect):
        """非对称: 政权向后(时间晚)延续便宜, 向前(时间早)追溯昂贵——
        「政权存续至继任者出现」"""
        r, c = cell
        r0, c0, r1, c1 = rect
        if r > r1:
            dr = (r - r1) * 3    # 标签之后 (延续)
        elif r < r0:
            dr = (r0 - r) * 10   # 标签之前 (追溯)
        else:
            dr = 0
        dc = max(c0 - c, c - c1, 0)
        return dr + dc

    raw_polities = []  # dict(name, color, cells, seeds, raws)
    multi_split_report = []
    unnamed_report = []

    for cid, comp in enumerate(comps):
        slist = comp_seeds.get(cid, [])
        groups = defaultdict(list)
        for s in slist:
            groups[s["name"]].append(s)
        if not groups:
            r0 = min(r for r, _ in comp["cells"])
            unnamed_report.append(
                (comp["color"], year_of(r0), len(comp["cells"]),
                 sorted(comp["cells"])[0]))
            if len(comp["cells"]) <= 1:
                continue  # 孤立杂点, 丢弃
            raw_polities.append(dict(name="", color=comp["color"],
                                     cells=comp["cells"], seeds=[], raws=set()))
        elif len(groups) == 1:
            name = next(iter(groups))
            raw_polities.append(dict(name=name, color=comp["color"],
                                     cells=comp["cells"], seeds=slist,
                                     raws={s["raw"] for s in slist}))
        else:
            assign = defaultdict(set)
            gitems = sorted(groups.items(),
                            key=lambda kv: min(s["rect"][0] for s in kv[1]))
            for cell in comp["cells"]:
                best, bestd = None, None
                for gname, gseeds in gitems:
                    d = min(rect_dist(cell, s["rect"]) for s in gseeds)
                    if bestd is None or d < bestd:
                        best, bestd = gname, d
                assign[best].add(cell)
            multi_split_report.append(
                (comp["color"], [g for g, _ in gitems],
                 year_of(min(r for r, _ in comp["cells"]))))
            for gname, gseeds in gitems:
                if assign[gname]:
                    raw_polities.append(dict(
                        name=gname, color=comp["color"], cells=assign[gname],
                        seeds=gseeds, raws={s["raw"] for s in gseeds}))

    for s in nofill_seeds:
        raw_polities.append(dict(name=s["name"], color=None,
                                 cells=set(rect_cells(s["rect"])),
                                 seeds=[s], raws={s["raw"]}))

    # --- 同名同色政权跨块合并 (如 秦 / 燕 / 西域都护府 / 西辽 的多段) ---
    merged = {}
    for p in raw_polities:
        key = (p["name"], p["color"]) if p["name"] else id(p)
        if key in merged:
            merged[key]["cells"] |= p["cells"]
            merged[key]["seeds"] += p["seeds"]
            merged[key]["raws"] |= p["raws"]
        else:
            merged[key] = p
    polities = list(merged.values())

    # --- 批注 ---
    notes = {}
    try:
        import zipfile
        z = zipfile.ZipFile(XLSX)
        xml = z.read("xl/comments1.xml").decode("utf-8")
        for m in re.finditer(r'<comment ref="([^"]+)"[^>]*>(.*?)</comment>', xml, re.S):
            ref, body = m.groups()
            txt = "".join(re.findall(r"<t[^>]*>([^<]*)</t>", body))
            txt = re.sub(r"^文西:\s*", "", txt).strip()
            mm = re.match(r"([A-Z]+)(\d+)", ref)
            ccol = 0
            for ch in mm.group(1):
                ccol = ccol * 26 + ord(ch) - 64
            notes[(int(mm.group(2)), ccol)] = txt
    except KeyError:
        pass
    for p in polities:
        for cell, txt in notes.items():
            if cell in p["cells"]:
                p["note"] = p.get("note", "") + txt

    # --- 简介 / 精确年份 ---
    try:
        sys.path.insert(0, str(ROOT / "tools"))
        from intros import INTROS
    except ImportError:
        INTROS = {}
    try:
        from years import EXACT
    except ImportError:
        EXACT = {}
    exact_used = set()

    # --- 输出 ---
    def runs_of(cells):
        by_row = defaultdict(list)
        for r, c in cells:
            by_row[r].append(c)
        runs = []
        for r in sorted(by_row):
            cols = sorted(by_row[r])
            start = prev = cols[0]
            for c in cols[1:]:
                if c == prev + 1:
                    prev = c
                    continue
                runs.append([r - MIN_ROW, start - MIN_COL, prev - MIN_COL])
                start = prev = c
            runs.append([r - MIN_ROW, start - MIN_COL, prev - MIN_COL])
        return runs

    def merge_label_rects(rects):
        """相邻/相接的同名标签矩形合并为包围盒 (跨格拆分名字拼回)"""
        rects = [list(x) for x in rects]
        changed = True
        while changed:
            changed = False
            for i in range(len(rects)):
                for j in range(i + 1, len(rects)):
                    a, b = rects[i], rects[j]
                    if (a[0] <= b[2] + 1 and b[0] <= a[2] + 1 and
                            a[1] <= b[3] + 1 and b[1] <= a[3] + 1):
                        rects[i] = [min(a[0], b[0]), min(a[1], b[1]),
                                    max(a[2], b[2]), max(a[3], b[3])]
                        rects.pop(j)
                        changed = True
                        break
                if changed:
                    break
        return rects

    max_used_row = max(r for r, _ in color) if color else MAX_ROW
    for p in polities:
        max_used_row = max(max_used_row, max(r for r, _ in p["cells"]))

    out_pol = []
    for p in sorted(polities, key=lambda p: (min(r for r, _ in p["cells"]),
                                             min(c for _, c in p["cells"]))):
        rows = [r for r, _ in p["cells"]]
        cols = [c for _, c in p["cells"]]
        y0, y1 = year_of(min(rows)), year_of(max(rows)) + YEAR_STEP
        regs = []
        for rn, rc0, rc1 in REGIONS:
            if any(rc0 <= c <= rc1 for c in cols):
                regs.append(rn)
        lrects = [(s["rect"][0], s["rect"][1], s["rect"][2], s["rect"][3])
                  for s in p["seeds"] if not s["virtual"]]
        lrects = merge_label_rects(lrects) if lrects else []
        labels = [[r0 - MIN_ROW, c0 - MIN_COL, r1 - MIN_ROW, c1 - MIN_COL]
                  for r0, c0, r1, c1 in lrects]
        name = p["name"]
        intro = INTROS.get(f"{name}@{y0}") or INTROS.get(name) or ""
        item = dict(id=len(out_pol), name=name, color=p["color"],
                    y0=y0, y1=y1, regions=regs,
                    runs=runs_of(p["cells"]), labels=labels)
        for k in (f"{name}@{y0}", name):
            if k in EXACT:
                item["ex"] = list(EXACT[k])
                exact_used.add(k)
                break
        raws = {r for r in p["raws"] if r != name and r != "（切分锚点）"}
        if raws:
            item["raw"] = "、".join(sorted(raws))
        if p.get("note"):
            item["note"] = p["note"]
        if intro:
            item["intro"] = intro
        out_pol.append(item)

    data = dict(
        meta=dict(yearStart=YEAR_START, yearStep=YEAR_STEP,
                  rows=max_used_row - MIN_ROW + 1, cols=MAX_COL - MIN_COL + 1,
                  source="朝代跨度表.xlsx / Sheet4"),
        regions=[dict(name=n, c0=c0 - MIN_COL, c1=c1 - MIN_COL)
                 for n, c0, c1 in REGIONS],
        polities=out_pol,
    )
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    (ROOT / "docs" / "data.json").write_text(payload, encoding="utf-8")
    # data.js: 使 file:// 直接打开 index.html 也能加载数据
    (ROOT / "docs" / "data.js").write_text(
        "window.DYNASTY_DATA=" + payload + ";", encoding="utf-8")

    # --- 审核报告 ---
    lines = ["# 解析审核报告\n"]
    lines.append(f"- 政权总数: **{len(out_pol)}**（其中未标注色块 {len(unnamed_report)} 个）")
    lines.append(f"- 简介覆盖: {sum(1 for p in out_pol if p.get('intro'))} 条\n")

    lines.append("## 名称修正清单（请审核）\n")
    lines.append("| 原文 | 修正为 | 位置 |")
    lines.append("|---|---|---|")
    seen = set()
    for raw, new, coord in sorted(renames_used, key=lambda x: x[0]):
        if (raw, new) in seen:
            continue
        seen.add((raw, new))
        lines.append(f"| {raw} | {new} | {coord} |")

    lines.append("\n## 多标签色块的切分（按最近标签划归格子）\n")
    for colr, names, y in sorted(multi_split_report, key=lambda x: x[2]):
        lines.append(f"- {y} 年起 {colr}: {' / '.join(names)}")

    lines.append("\n## 未标注的色块（原表无名字）\n")
    for colr, y, n, cell in sorted(unnamed_report, key=lambda x: x[1]):
        lines.append(f"- {y} 年起 {colr}, {n} 格, 起于 "
                     f"{get_column_letter(cell[1])}{cell[0]}")

    lines.append("\n## 全部政权（按开始年排序）\n")
    lines.append("| 起 | 止 | 名称 | 原文 | 颜色 | 区域 | 简介 |")
    lines.append("|---|---|---|---|---|---|---|")
    for p in out_pol:
        lines.append(f"| {p['y0']} | {p['y1']} | {p['name'] or '（未标注）'} "
                     f"| {p.get('raw', '')} | {p['color'] or '无填充'} "
                     f"| {'/'.join(p['regions'])} | {'✓' if p.get('intro') else ''} |")

    (ROOT / "tools" / "report.md").write_text("\n".join(lines), encoding="utf-8")
    unused_exact = [k for k in EXACT if k not in exact_used]
    if unused_exact:
        print("WARN 未匹配的精确年份键:", unused_exact)
    print(f"OK: {len(out_pol)} polities -> docs/data.json "
          f"(精确年份 {sum(1 for p in out_pol if 'ex' in p)} 条)")
    print(f"report -> tools/report.md")


if __name__ == "__main__":
    main()
